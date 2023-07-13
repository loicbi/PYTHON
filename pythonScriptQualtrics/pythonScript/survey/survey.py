import requests
import zipfile
import io, os, re, platform
import json
import pandas as pd
# from azure.storage.blob import BlobServiceClient, BlobClient
import logging
from azure.storage.blob import (BlobServiceClient, BlobClient, ContainerClient)
from datetime import datetime
import openpyxl
from openpyxl.utils.cell import get_column_letter

# datetime object containing current date and time
now = datetime.now()

# Setting user Parameters
api_token = "v8D9Zv0Jlyg84pXR8fgqMiRe0MnuXMsbG1fYWEvy"
file_format = "csv"
data_center = 'yul1'
survey_list = ['SV_5hjhS6FaaBCyl6K','SV_2mMFEB1OKln3F9I',  'SV_byHLBVUjHMAz6Qu', 'SV_8CagbMF0wOqQHbw', 'SV_8cVD0wbwDtZPoHk',
               'SV_1ZkxR7W55aTgpdY', 'SV_4NPoN9SryGI8hb8', 'SV_9QOf8rrIQ7u3tk2', 'SV_egHDV1AGAz2O28e', 'SV_cGwgMDMXAzDyRVk',
               'SV_a4V4YTON9bVlnaC','SV_0eLB18rACdf7bH8', 'SV_8jC7Blg2OEmwhp4', 'SV_3zcdvGe3IpCg6EK',
               'SV_0TedVXikBFGuTau', 'SV_0AjwVuT4XBOG0dw', 'SV_42BUj5ho4V6fGnQ', 'SV_6FNzf9P7QmZSeO2', 'SV_eaGvg6mDwYNwjwq',
               'SV_7PZnjqEOb5lBAlU']
# BAD == 'SV_5hjhS6FaaBCyl6K',
#
# survey_list = ['SV_42BUj5ho4V6fGnQ', 'SV_8jC7Blg2OEmwhp4']

# survey_id = 'SV_egHDV1AGAz2O28e'  # SV_egHDV1AGAz2O28e SV_2mMFEB1OKln3F9I  SV_9QOf8rrIQ7u3tk2 SV_6S5ZqosmZgtbTYG SV_4NPoN9SryGI8hb8 SV_8jC7Blg2OEmwhp4(not response)
# SV_7PZnjqEOb5lBAlU

survey_list_not_reponse = []
for survey_id in survey_list:

    # Setting static parameters
    request_check_progress = 0
    progress_status = "in progress"

    # post
    startDate = '2010-01-01T00:00:00Z'
    endDate = '2050-12-31T00:00:00Z'
    payload = {"format": "json",
               "startDate": startDate,
               "endDate": endDate,
               "compress": False, }
    headers = {
        "content-type": "application/json",
        "x-api-token": api_token,
    }
    get_FileId = ''


    # get url realTime
    def get_url_completed():
        url = f"https://yul1.qualtrics.com/API/v3/surveys/{survey_id}/export-responses"
        # Step 1: Creating Data Export
        get_ProgressId = requests.request("POST", url, json=payload, headers=headers).json()['result']['progressId']
        print(f"Progress Id is:  {get_ProgressId}")
        # https://yul1.qualtrics.com/API/v3/surveys/SV_2mMFEB1OKln3F9I/export-responses/ES_6kWBjbYTzJmEDEW
        url_progressId = f"https://yul1.qualtrics.com/API/v3/surveys/{survey_id}/export-responses/{get_ProgressId}"

        request_check_progress = 0.0
        progress_status = "in progress"
        get_FileId = dict()
        # Step 2: Checking on Data Export Progress and waiting until export is ready
        while request_check_progress < 100 and progress_status != "complete":
            request_check_response = requests.request("GET", url_progressId, headers=headers)
            request_check_progress = request_check_response.json()["result"]["percentComplete"]
            get_FileId = request_check_response.json()
        get_FileId = get_FileId["result"]['fileId']

        url_complete = f"https://{data_center}.qualtrics.com/API/v3/surveys/{survey_id}/export-responses/{get_FileId}/file"
        get_responses_surveys_url = url_complete
        return get_responses_surveys_url


    get_responses_surveys_url = get_url_completed()
    # get_responses_surveys_url = 'https://ca1.qualtrics.com/API/v3/surveys/SV_2mMFEB1OKln3F9I/export-responses/ebba1a47-082d-4d55-9093-dc5af04ec15c-def/file'
    """Get surveyname """
    url = f"https://yul1.qualtrics.com/API/v3/surveys/{survey_id}"

    """GET SURVEY NAME """

    response = requests.request("GET", url, headers=headers)
    """ERROR 404"""
    if '404' in response.json()['meta']['httpStatus']:
        print(
            f"SURVEY NAME: {survey_id} HAS BEEN REMOVED, check here ==> https://api.qualtrics.com/73d7e07ec68b2-get-survey")
        surveyName = f"SURVEY_NAME_GENERATED_{survey_id}"
        # continue
    else:
        """OK 200"""
        surveyName = response.json()['result']['name']
        # surveyName = "                     @@@ gdgyhd dggbdyhd - hgdyjhd(246152115-12-44                  )"  # test
        surveyName = re.sub(r"[^A-Za-z0-9]", " ", surveyName).rstrip().lstrip().replace("  ", "").replace(' ',
                                                                                                          '_').upper()
    print(f"SurveyId is: {survey_id}")
    print(f"SurveyName is: {surveyName}")

    """GET QUESTIONS """


    def get_all_questions():
        url_question = f"https://yul1.qualtrics.com/API/v3/survey-definitions/{survey_id}/questions"
        response = requests.request("GET", url_question, headers=headers)
        list_1 = response.json()['result']['elements']
        # number of question
        number_question_by_number = len(response.json()['result']['elements'])
        # print(number_question_by_number)
        question_id_single_without_choice = []
        # _old_question_id_array_with_choice = []
        question_id_array_with_choice = []
        # for i in range(number_question_by_number):
        #     # print(list_1[i]['QuestionType'])
        #
        #     try:
        #         # single question DB WITH ChoiceOrder
        #         if (list_1[i]['QuestionType'] == 'DB') and (list_1[i]['ChoiceOrder'] == []):
        #             question_id_single_without_choice.append(list_1[i]['QuestionID'])
        #         else:
        #             # MULTIPLE CHOICE
        #             if (len(list_1[i]['ChoiceOrder']) > 0):
        #                 for j in range(len(list_1[i]['ChoiceOrder'])):
        #                     if list_1[j]['QuestionID'] not in question_id_single_without_choice:
        #                         # print(list_1[j]['QuestionID'], '###### >', list_1[j]['QuestionType'],
        #                         #       len(list_1[j]['ChoiceOrder']))
        #                         question_id_array_with_choice.append(f"{list_1[i]['QuestionID']}_{j + 1}")
        #     except KeyError:
        #         # SINGLE QUESTION TE WITHOUT ChoiceOrder
        #         if (list_1[i]['QuestionType'] == 'TE' and (
        #                 list_1[i]['QuestionID'] not in question_id_single_without_choice + question_id_array_with_choice)):
        #             # print(list_1[i]['QuestionID'], '@@@>', list_1[i]['QuestionType'])
        #             question_id_single_without_choice.append(list_1[i]['QuestionID'])
        #         # print(list_1[i]['QuestionID'], ':::::::>', list_1[i]['QuestionType'])

        for i in range(number_question_by_number):
            # CHECK IF ChoiceOrder EXISTS
            try:
                if 'ChoiceOrder' not in list_1[i].keys():
                    # print(f"{list_1[i]['QuestionID']} ::: {list_1[i]['QuestionType']} ::: NA ")
                    question_id_single_without_choice.append(list_1[i]['QuestionID'])
                else:
                    # print(f"{list_1[i]['QuestionID']} ::: {list_1[i]['QuestionType']} has Choice order => {list_1[i]['ChoiceOrder']} ")
                    if list_1[i]['ChoiceOrder'] == []:
                        question_id_single_without_choice.append(list_1[i]['QuestionID'])
                    else:
                        """Multiple questions so increment _n"""
                        # print(f"{list_1[i]['QuestionID']} ::QuestionType:: {list_1[i]['Choices']} ::ChoiceOrder:: => {list_1[i]['ChoiceOrder']} ")
                        # if list_1[i]['QuestionID'] != 'QID13':
                        if list_1[i]['QuestionID']:
                            # print(f"{list_1[i]['QuestionID']} ::: {list_1[i]['QuestionType']} has Choice order => {list_1[i]['ChoiceOrder']} ")
                            for j in range(len(list_1[i]['ChoiceOrder'])):
                                # print(f"{list_1[i]['QuestionID']}_{j + 1}")
                                #     print(list_1[j]['QuestionID'], '###### >', list_1[j]['QuestionType'], len(list_1[j]['ChoiceOrder']))
                                question_id_array_with_choice.append(f"{list_1[i]['QuestionID']}_{j + 1}")
                        # else:
                        #     print(
                        #         f"{list_1[i]['QuestionID']} ::: {list_1[i]['QuestionType']} has Choice order => {list_1[i]['ChoiceOrder']} ")
                        #     for j in range(len(list_1[i]['ChoiceOrder'])):
                        #         question_id_array_with_choice.append(f"{list_1[i]['QuestionID']}_{j + 1}")


            except KeyError:
                print(list_1[i]['QuestionID'], '@@@@@@@@@@@@@@@@@@@@@@@@ERROR@@@@@@@@@@@@@@@@@@@@@@@@',
                      list_1[i]['QuestionType'])

        # print(question_id_single_without_choice)
        questions = question_id_single_without_choice + question_id_array_with_choice
        return questions


    allQuesQualtricsApi = get_all_questions()
    download_request_response = requests.get(get_responses_surveys_url, headers=headers)

    download_request_response = download_request_response.json()
    # print(download_request_response)
    # str to json or dict
    number_responses_id = len(download_request_response['responses'])
    print(f"Number Response ID: {number_responses_id}")
    """(A) ===> add responseId to ['responses'][0]['values'] to get all fields's request"""

    dict_all_responses = {}
    questions_list_responses = []
    if number_responses_id > 0:
        download_request_response['responses'][0]['values']['responseId'] = download_request_response['responses'][0][
            'responseId']
        for i in range(number_responses_id):
            """ (A) """
            download_request_response['responses'][i]['values']['responseId'] = \
                download_request_response['responses'][i][
                    'responseId']
            dict_all_responses[f"responseId_of_{download_request_response['responses'][i]['responseId']}"] = \
                download_request_response['responses'][i]['values']
            """ Add questions lists from response """
            for questionItem in list(
                    dict_all_responses[f"responseId_of_{download_request_response['responses'][i]['responseId']}"]):
                questions_list_responses.append(questionItem)

            item_value = dict_all_responses[f"responseId_of_{download_request_response['responses'][i]['responseId']}"]

            for x in list(item_value.keys()):
                if item_value[x] == []:
                    item_value[x] = ['']
    else:
        print("Not response")
        survey_list_not_reponse.append(survey_id)
        print("############FINISH#############".center(100))
        continue
        # exit()

    # remove double
    questions_list_responses = list(dict.fromkeys(questions_list_responses))
    # print('debut', len(questions_list_responses))
    questionsToAdd = list(set(allQuesQualtricsApi) - set(questions_list_responses))
    # print('a ajouter ', len(questionsToAdd))
    toAdd = []
    for a in questionsToAdd:
        questions_list_responses.append(a)

    # add questions not exists to send to snowflake
    for question in questions_list_responses:
        for x in list(dict_all_responses.keys()):
            if question not in list(dict_all_responses[x].keys()):
                # print(f"{question}")
                dict_all_responses[x][question] = ['']

    print(f"Number question is: {len(questions_list_responses)}")
    """ Set header as column and macth with keys  """

    # print(dict_all_responses)
    # dataframe = pd.DataFrame(dict_all_responses['responseId_of_R_3QFOuZrWGeASxuI'])

    a = []
    for x in list(dict_all_responses.keys()):
        """CONVERT STR TO DATETIME USING TIMEZONE """
        dict_all_responses[x]['startDate'] = datetime.strptime(dict_all_responses[x]['startDate'], "%Y-%m-%dT%H:%M:%SZ")
        dict_all_responses[x]['endDate'] = datetime.strptime(dict_all_responses[x]['endDate'], "%Y-%m-%dT%H:%M:%SZ")
        # print(dict_all_responses[x]) # R_1CrJf5E3zyTYgwz
        # dataframe = pd.DataFrame(dict_all_responses[x])
        # a.append(dataframe)
        dataframe = pd.DataFrame.from_dict(dict_all_responses[x], orient='index').transpose()
        a.append(dataframe)

    # new_dataframe = pd.concat(a).drop_duplicates()
    new_dataframe = pd.concat(a)
    # print(new_dataframe.dtypes)
    # exit()
    print(f"Number Row is: {new_dataframe.shape[0]}")
    path_system = os.getcwd()
    print(f"path_system is : {path_system}")
    new_dataframe.to_excel(f'{path_system}/{survey_id}.xlsx', sheet_name=survey_id)
    # new_dataframe.to_excel(fr'E:\\Bitbucket\\andreloic\\pythonScript\\{survey_id}.xlsx', sheet_name=survey_id)
    # new_dataframe.to_json(path_or_buf="E:\\Bitbucket\\andreloic\\pythonScript\\data.json", orient='split')

    """REPLACE VALUE [''] by '' """
    wb = openpyxl.load_workbook(f'{path_system}/{survey_id}.xlsx')
    wb.sheetnames
    sheet = wb[survey_id]
    number_rows = sheet.max_row
    number_columns = sheet.max_column
    replacement = {"['']": '', }
    for i in range(number_columns):
        for k in range(number_rows):
            cell = str(sheet[get_column_letter(i + 1) + str(k + 1)].value)
            for key in replacement.keys():
                if str(cell) == key:
                    newCell = replacement.get(key)
                    sheet[get_column_letter(i + 1) + str(k + 1)] = str(newCell)
    wb.save(f'{path_system}/{survey_id}.xlsx')
    # exit()
    print(f'{path_system}/{survey_id}.xlsx had been updated')

    """ AZURE BLOB STORAGE """

    storage_account_key = "3DRT9Nrt9Ab8xKZNW8A4PDcf2H9gg4IyLpqD7j1SaSIzhhKSuqdgAmKH9xLE2PpPf/ALz8fWL4Vsmk3PC0NyXw=="
    storage_account_name = "samtdaentcadl"
    container_name = "dev-dl"
    connection_string = "DefaultEndpointsProtocol=https;AccountName=samtdaentcadl;AccountKey=3DRT9Nrt9Ab8xKZNW8A4PDcf2H9gg4IyLpqD7j1SaSIzhhKSuqdgAmKH9xLE2PpPf/ALz8fWL4Vsmk3PC0NyXw==;EndpointSuffix=core.windows.net"
    file_path_blob = f"ca/THIRD_PARTY_DATA/QUALTRICS/{surveyName}/"


    # https://thats-it-code.com/azure/azure-blob-storage-operation-using-python/
    def uploadToBlobStorage(file_path, file_name):
        blob_service_client = BlobServiceClient.from_connection_string(connection_string)
        blob_client = blob_service_client.get_blob_client(container=container_name, blob=file_name)
        # Get the blob client to be deleted
        todelete_blob_client = blob_service_client.get_blob_client(container=container_name,
                                                                   blob=f"{file_path_blob}{survey_id}.xlsx")

        # print(f'todelete_blob_client exists ? : {todelete_blob_client.exists()}')
        # print(f'todelete_blob_client: {todelete_blob_client}')

        """PYTHON 3.10.4 """
        container = ContainerClient.from_connection_string(conn_str=connection_string, container_name=container_name,
                                                           credential=storage_account_key)
        """PYTHON 3.6.4 """
        # container = ContainerClient.from_connection_string(conn_str=connection_string, container=container_name)
        blob_list = container.list_blobs()
        # for blob in blob_list:
        #     print(blob.name + '\n')
        # https://pypi.org/project/azure-storage-blob/
        with open(file_path, "rb") as data:
            # check list bob
            """METHOD 1"""
            for blob in blob_list:
                try:
                    # Check if the blob exists
                    if blob.name == file_name:
                        # Delete blob
                        todelete_blob_client.delete_blob()
                        print("Blob deleted successfully!")
                    #     blob_client.upload_blob(data)
                except Exception as e:
                    print("Failed to delete blob. Error:" + str(e))
            blob_client.upload_blob(data)
            print(f"Uploaded {file_name}.")

            """METHOD 2"""
            # try:
            # # Check if the blob exists
            #     if todelete_blob_client.exists():
            #         # Delete blob
            #         todelete_blob_client.delete_blob()
            #         print("Blob deleted successfully!")
            #     else:
            #         print("Blob does not exist!")
            #     #     blob_client.upload_blob(data)
            # except Exception as e:
            #     print("Failed to delete blob. Error:" + str(e))


    # calling a function to perform upload
    uploadToBlobStorage(f'{path_system}/{survey_id}.xlsx',
                        f"ca/THIRD_PARTY_DATA/QUALTRICS/{surveyName}/{survey_id}.xlsx")
    # uploadToBlobStorage(f"E:\\Bitbucket\\andreloic\\pythonScript\\{survey_id}.xlsx",
    #                     f"ca\THIRD_PARTY_DATA\QUALTRICS\{surveyName}\{survey_id}__{now.strftime('%d%m%Y_%H:%M:%S')}.xlsx")

    print("############FINISH#############".center(100))
print(f"survey_list_not_reponse: {survey_list_not_reponse}")
