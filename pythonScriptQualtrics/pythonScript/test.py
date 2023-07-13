import requests
#
# # url = "https://httpbin.org/post"
# url = "https://ca1.qualtrics.com/API/v3/surveys/SV_2mMFEB1OKln3F9I/export-responses/413e5919-9abb-46df-b4d7-adf8ad5032da-def/file"
#
# data = {
# 	"id": 1001,
# 	"name": "gerek",
# 	"passion": "coding",
# }
#
# headers = {
#     "content-type": "application/json",
#     "x-api-token": 'v8D9Zv0Jlyg84pXR8fgqMiRe0MnuXMsbG1fYWEvy',
# }
#
# # response = requests.post(url, json=data)
# response = requests.get(url, headers=headers)
#
# # print(response.text)
# print("Status Code", response.status_code)
# # print("JSON Response ", response.json())
#
import tablib

data = tablib.Dataset(headers=['First Name', 'Last Name', 'Age'])
for i in [('Kenneth', 'Reitz', 22), ('Bessie', 'Monke', 21)]:
    data.append(i)

print(data.export('json'))

print(data.export('yaml'))
#
data.export('xlsx')
print(type([('Kenneth', 'Reitz', 22), ('Bessie', 'Monke', 21)]))
with open('output.xlsx', 'wb') as f:
    f.write(data.export('xlsx'))

#
# data.export('df')

import pandas as pd

print("#########################".center(100))

#
# dict = {'abc':'1/2/2003', 'def':'1/5/2017', 'ghi':'4/10/2013'}
#
# df["Date"] = df["Member"].apply(lambda x: d.get(x))


# dataframe1 = pd.DataFrame({'columnA': ['fff'],
#                            'columnB': 20})
#
# dataframe2 = pd.DataFrame({'columnA': ['d'],
#                            'columnB': 20})
#
# # Concatenating dataframes without duplicates
# new_dataframe = pd.concat([dataframe1, dataframe2]).drop_duplicates()
# print([dataframe1, dataframe2])
# Display concatenated dataframe
# print(new_dataframe)


# print(list(data_1['startDate'].split()))
# for key in data_1:
#     if isinstance(data_1[key], str):
#         print(key, '->', list(data_1[key].split(',')))

# else:`
#     [int(data_1[key]) for i in str(12345)]

data_survey = {'startDate': '2023-03-31 15:31:53', 'endDate': '2023-03-31 15:35:21', 'status': 0,
               'ipAddress': '216.211.50.3', 'progress': 100, 'duration': 208, 'finished': 1,
               'recordedDate': '2023-03-31T15:35:23.446Z', '_recordId': 'R_6VfRz3g8gHFfZwl',
               'recipientLastName': 'Silverman', 'recipientFirstName': 'Mitch',
               'recipientEmail': 'syndicatepharmacy@hotmail.com', 'externalDataReference': '0031R00002moFpqQAE',
               'locationLatitude': '48.3987', 'locationLongitude': '-89.3168', 'distributionChannel': 'email',
               'userLanguage': 'EN', 'QID2_NPS_GROUP': 2, 'QID2': 8, 'QID3_TEXT': 'next day delivery', 'QID4_1': 8,
               'QID4_2': 7, 'QID4_3': 8, 'QID4_4': 8, 'QID4_5': 9, 'QID4_6': 8, 'QID4_7': 8, 'QID4_8': 7, 'QID4_9': 7,
               'QID30_1': 7, 'QID9_1': 7, 'QID9_2': 6, 'QID9_3': 7, 'QID9_4': 6, 'QID9_5': 6, 'QID9_6': 6,
               'QID25': [''], 'Audience': '3', 'AccountBannerSAP_c': 'BG_PHA_RRX',
               'AccountCustomerGroupDescription': 'Not FHCP/ARP', 'AccountMBARateStructureCode': 'RRXEON',
               'AccountMBARateStructureDescription': "REMEDY'S RX ONT - UNB 14/28",
               'AccountOwner': 'SELVI RUBIO-NARANJO', 'AccountState': 'ON', 'sfAccountId': '0013600001RZ3bJAAT',
               'sfContactId': '0031R00002moFpqQAE',
               'Q_URL': 'https://mckessoncanada.qualtrics.com/jfe/form/SV_2mMFEB1OKln3F9I?Q_DL=OCJMZ1uqIhLqqxF_2mMFEB1OKln3F9I_CGC_ovRk9hoqpIbjhhz&Q_CHL=email',
               'AccountCarrierCode': 'GDW', 'AccountCarrierDescription': 'GARDEWINE', 'AccountDeliveryRoute': 'O2',
               'AccountMainDC': '450', 'sfAccount': 'SYNDICATE PHARMACY (Location)', 'sfPhone': '(807) 623-8844',
               'AccountMBAnumber': '456525', 'SAPMaster': '7001038', 'ContactType': 'Owner', 'Environment': 'Prod',
               'Complete': '3', 'BannerFN': 'Remedy’s Rx', 'FOLLOWUP_NEW': '0', 'responseId': 'R_6VfRz3g8gHFfZwl',
               'ContactPostalCode': [''], 'sfMailingState': [''], 'sfTitle': [''], 'QID24_TEXT': [''],
               'Additionalcomments': [''], 'Followup': [''], 'FullName': [''], 'Phone': [''], 'QID31_1': [''],
               'QID11_1': [''], 'QID11_2': [''], 'QID11_3': [''], 'QID11_4': [''], 'QID11_6': [''], 'QID11_7': [''],
               'QID28_1': [''], 'QID6_1': [''], 'QID6_2': [''], 'QID6_3': [''], 'QID6_4': [''], 'QID15_TEXT': [''],
               'QID17': [''], 'QID16': [''], 'QID18': [''], 'QID12_TEXT': [''], 'QID13': [''], 'sfName': [''],
               'sfOrgCompany': [''], 'sfSurveyID': [''], 'QID20': [''], 'QID21': [''], 'QID20_24_TEXT': [''],
               'QID21_9_TEXT': [''], 'sfMailingCity': [''], 'sfMailingCountry': [''], 'sfMailingPostalCode': [''],
               'QID18_9_TEXT': [''], 'QID14_1': [''], 'QID14_2': [''], 'QID14_3': [''], 'QID14_4': [''],
               'QID16_6_TEXT': [''], 'QID20_15': [''], 'QID18_2': [''], 'QID16_2': [''], 'QID20_7': [''],
               'QID2_11': [''], 'QID2_10': [''], 'QID2_2': [''], 'QID11_5': [''], 'QID20_11': [''], 'QID17_10': [''],
               'QID21_1': [''], 'QID20_17': [''], 'QID18_3': [''], 'QID20_6': [''], 'QID21_6': [''], 'QID5_1': [''],
               'QID2_3': [''], 'QID20_3': [''], 'QID20_20': [''], 'QID26_3': [''], 'QID20_2': [''], 'QID3': [''],
               'QID27': [''], 'QID29_3': [''], 'QID20_12': [''], 'QID7': [''], 'QID2_4': [''], 'QID1': [''],
               'QID10_1': [''], 'QID20_18': [''], 'QID21_3': [''], 'QID20_9': [''], 'QID18_5': [''], 'QID17_2': [''],
               'QID20_16': [''], 'QID17_6': [''], 'QID17_11': [''], 'QID16_3': [''], 'QID18_4': [''], 'QID20_21': [''],
               'QID17_12': [''], 'QID24': [''], 'QID20_4': [''], 'QID2_1': [''], 'QID29_2': [''], 'QID13_1': [''],
               'QID2_7': [''], 'QID20_10': [''], 'QID17_8': [''], 'QID32_2': [''], 'QID32_1': [''], 'QID2_8': [''],
               'QID15': [''], 'QID17_4': [''], 'QID26_1': [''], 'QID29_1': [''], 'QID21_4': [''], 'QID17_9': [''],
               'QID18_6': [''], 'QID12': [''], 'QID20_8': [''], 'QID8_1': [''], 'QID20_5': [''], 'QID2_9': [''],
               'QID25_1': [''], 'QID2_6': [''], 'QID32_3': [''], 'QID20_14': [''], 'QID17_13': [''], 'QID21_5': [''],
               'QID16_1': [''], 'QID17_1': [''], 'QID17_3': [''], 'QID17_7': [''], 'QID18_1': [''], 'QID20_13': [''],
               'QID20_19': [''], 'QID17_5': [''], 'QID21_2': [''], 'QID20_1': [''], 'QID26_4': [''], 'QID26_2': [''],
               'QID2_5': ['']}

data1 = {'startDate': '2023-03-24 16:27:41', 'endDate': '2023-03-24 16:27:49', 'status': 0,
         'ipAddress': '134.238.147.126', 'progress': 100, 'duration': 7, 'finished': 1,
         'recordedDate': '2023-03-24T16:27:49.594Z', '_recordId': 'R_1JQM1vVW7aQ1RNV', 'locationLatitude': '43.6922',
         'locationLongitude': '-79.4309', 'distributionChannel': 'anonymous', 'userLanguage': 'EN', 'QID2': 5,
         'responseId': 'R_1JQM1vVW7aQ1RNV', 'QID3_1': [''], 'QID3_2': [''], 'QID3_3': [''], 'QID16_1': [''],
         'QID16_2': [''], 'QID16_3': [''], 'QID17_FILE_ID': [''], 'QID17_FILE_NAME': [''], 'QID17_FILE_SIZE': [''],
         'QID17_FILE_TYPE': [''], 'QID18#1_1': [''], 'QID18#1_2': [''], 'QID18#1_3': [''], 'QID18#2_1': [''],
         'QID18#2_2': [''], 'QID18#2_3': [''], 'QID19_2': [''], 'QID19_3': [''], 'QID19_1': [''], 'QID20_BROWSER': [''],
         'QID20_VERSION': [''], 'QID20_OS': [''], 'QID20_RESOLUTION': [''], 'QID21_0_GROUP': [''],
         'QID21_1_GROUP': [''], 'QID21_2_GROUP': [''], 'QID21_G0_2_RANK': [''], 'QID21_G1_1_RANK': [''],
         'QID21_G2_3_RANK': [''], 'QID21_G0_1_RANK': [''], 'QID21_G1_2_RANK': [''], 'QID21_G1_3_RANK': [''],
         'QID26_FIRST_CLICK': [''], 'QID26_LAST_CLICK': [''], 'QID26_PAGE_SUBMIT': [''], 'QID26_CLICK_COUNT': [''],
         'QID27': [''], 'QID28_1': [''], 'QID28_2': [''], 'QID28_3': [''], 'QID21_G2_2_RANK': [''], 'QID17': [''],
         'QID2_4': [''], 'QID23': [''], 'QID18_3': [''], 'QID1_3': [''], 'QID15_3': [''], 'QID20': [''], 'QID1_2': [''],
         'QID24': [''], 'QID18_2': [''], 'QID2_3': [''], 'QID26': [''], 'QID15_2': [''], 'QID2_5': [''], 'QID1_4': [''],
         'QID21_3': [''], 'QID22_1': [''], 'QID2_2': [''], 'QID25': [''], 'QID21_1': [''], 'QID18_1': [''],
         'QID1_5': [''], 'QID22_3': [''], 'QID21_2': [''], 'QID2_1': [''], 'QID15_1': [''], 'QID22_2': [''],
         'QID1_1': ['']}

data2 = {'startDate': '2023-03-29 11:38:43', 'endDate': '2023-03-29 11:42:02', 'status': 1, 'progress': 100,
         'duration': 198, 'finished': 1, 'recordedDate': '2023-03-29T11:42:08.723Z', '_recordId': 'R_1CrJf5E3zyTYgwz',
         'locationLatitude': '45.5332', 'locationLongitude': '-73.6091', 'distributionChannel': 'preview',
         'userLanguage': 'EN', 'QID2': 3, 'QID3_1': 2, 'QID3_2': 2, 'QID3_3': 2, 'QID16_1': 'Yy', 'QID16_2': 'Trd',
         'QID16_3': 'Hfch', 'QID17_FILE_ID': 'F_2TZYMX8LBcVrAME',
         'QID17_FILE_NAME': 'E90D86A7-3AC8-4B89-A01B-C9E21C3A8400.jpeg', 'QID17_FILE_SIZE': 4685845,
         'QID17_FILE_TYPE': 'image/jpeg', 'QID18#1_1': 1, 'QID18#1_2': 1, 'QID18#1_3': 1, 'QID18#2_1': 1,
         'QID18#2_2': 2, 'QID18#2_3': 1, 'QID19_1': 2, 'QID19_2': 2, 'QID19_3': 1, 'QID20_BROWSER': 'Safari iPhone',
         'QID20_VERSION': '16.0', 'QID20_OS': 'iPhone', 'QID20_RESOLUTION': '390x844', 'QID21_0_GROUP': ['1'],
         # 'QID21_1_GROUP': ['3', '2'],
         'QID21_2_GROUP': [''], 'QID21_G0_1_RANK': 1, 'QID21_G1_2_RANK': 2,
         'QID21_G1_3_RANK': 1, 'QID26_FIRST_CLICK': 2.027, 'QID26_LAST_CLICK': 192.993, 'QID26_PAGE_SUBMIT': 193.705,
         'QID26_CLICK_COUNT': 109, 'QID27': 7, 'QID28_1': 5660, 'QID28_2': 65, 'QID28_3': 534453344477543,
         'responseId': 'R_1CrJf5E3zyTYgwz', 'ipAddress': [''], 'QID21_G0_2_RANK': [''], 'QID21_G1_1_RANK': [''],
         'QID21_G2_3_RANK': [''], 'QID21_G2_2_RANK': [''], 'QID2_4': [''], 'QID26': [''], 'QID20': [''], 'QID17': [''],
         'QID21_3': [''], 'QID22_3': [''], 'QID2_3': [''], 'QID24': [''], 'QID18_3': [''], 'QID23': [''],
         'QID2_5': [''], 'QID2_2': [''], 'QID1_3': [''], 'QID21_2': [''], 'QID18_1': [''], 'QID2_1': [''],
         'QID1_1': [''], 'QID1_2': [''], 'QID15_2': [''], 'QID22_1': [''], 'QID15_3': [''], 'QID1_5': [''],
         'QID18_2': [''], 'QID21_1': [''], 'QID25': [''], 'QID22_2': [''], 'QID1_4': [''], 'QID15_1': ['']
         }


# a = []
# df1 = pd.DataFrame.from_dict(data1, orient='index')
# df2 = pd.DataFrame.from_dict(data2, orient='index')
# df1 = df1.transpose()
# df2 = df2.transpose()
# new_dataframe = pd.concat([df1, df2])
# a.append(new_dataframe)
# print(a)


import openpyxl
from openpyxl.utils.cell import get_column_letter

# exit()

wb = openpyxl.load_workbook('E:\\Bitbucket\\andreloic\\pythonScript\\survey\\SV_5hjhS6FaaBCyl6K.xlsx')
wb.sheetnames
sheet = wb["SV_5hjhS6FaaBCyl6K"]
number_rows = sheet.max_row
number_columns = sheet.max_column

replacement = {"['']": '',}

for i in range(number_columns):
    for k in range(number_rows):
        cell = str(sheet[get_column_letter(i + 1) + str(k + 1)].value)
        for key in replacement.keys():
            if str(cell) == key:
                newCell = replacement.get(key)
                sheet[get_column_letter(i + 1) + str(k + 1)] = str(newCell)
wb.save('E:\\Bitbucket\\andreloic\\pythonScript\\survey\\SV_5hjhS6FaaBCyl6K.xlsx')



